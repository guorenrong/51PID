# importing Qt widgets
from msilib.schema import CheckBox
from pickle import TRUE
from tkinter import W
from PyQt5 import QtWidgets
from PyQt5 import QtCore
from PyQt5.QtWidgets import *
import sys
import random
from PyQt5.QtChart import QDateTimeAxis, QValueAxis, QSplineSeries, QChart, QChartView
from PyQt5.QtWidgets import QApplication
from PyQt5.QtGui import QPainter
from PyQt5.QtCore import QDateTime, Qt, QTimer
import serial
import struct
import threading
import time
import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.chart import Reference, LineChart
# importing pyqtgraph as pg
import pyqtgraph as pg
import queue
import SettingDialog
from PyQt5.QtGui import QIcon,QPixmap

class ChartView(QChartView, QChart):
    def __init__(self, *args, **kwargs):
        super(ChartView, self).__init__(*args, **kwargs)
        self.resize(1920, 1080)
        self.setRenderHint(QPainter.Antialiasing)  # 抗锯齿
        self.chart_init()
        #self.timer_init()

    def timer_init(self):
        # 使用QTimer，1秒触发一次，更新数据
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.drawLine)
        self.timer.start(1000)

    def chart_init(self):
        self.chart = QChart()
        self.series_1 = QSplineSeries()
        # 设置曲线名称
        self.series_1.setName("实时温度")
        # 把曲线添加到QChart的实例中
        self.chart.addSeries(self.series_1)
        #self.series_1.connect(Qt.MouseHovered)

        #self.series_2 = QSplineSeries()
        # 设置曲线名称
        #self.series_2.setName("实时数据-2")
        # 把曲线添加到QChart的实例中
        #self.chart.addSeries(self.series_2)

        # 声明并初始化X轴，Y轴
        self.dtaxisX = QDateTimeAxis()
        self.vlaxisY = QValueAxis()
        # 设置坐标轴显示范围
        self.dtaxisX.setMin(QDateTime.currentDateTime().addSecs(-300*1))
        self.dtaxisX.setMax(QDateTime.currentDateTime().addSecs(0))
        self.vlaxisY.setMin(0)
        self.vlaxisY.setMax(1500)
        self.series_1.setColor(Qt.GlobalColor.red)
        # 设置X轴时间样式
        self.dtaxisX.setFormat("hh:mm:ss")
        # 设置坐标轴上的格点
        self.dtaxisX.setTickCount(11)
        self.vlaxisY.setTickCount(11)
        # 设置坐标轴名称
        self.dtaxisX.setTitleText("时间")
        self.vlaxisY.setTitleText("温度值")
        # 设置网格不显示
        self.vlaxisY.setGridLineVisible(False)
        # 把坐标轴添加到chart中
        self.chart.addAxis(self.dtaxisX,Qt.AlignBottom)
        self.chart.addAxis(self.vlaxisY,Qt.AlignLeft)
        # 把曲线关联到坐标轴
        self.series_1.attachAxis(self.dtaxisX)
        self.series_1.attachAxis(self.vlaxisY)
        

        # 把曲线关联到坐标轴
        #self.series_2.attachAxis(self.dtaxisX)
        #self.series_2.attachAxis(self.vlaxisY)

        self.setChart(self.chart)

    def drawLine(self, yint):
        # 获取当前时间
        bjtime = QDateTime.currentDateTime()
        # 更新X轴坐标
        self.dtaxisX.setMin(QDateTime.currentDateTime().addSecs(-300*1))
        self.dtaxisX.setMax(QDateTime.currentDateTime().addSecs(0))
        # 当曲线上的点超出X轴的范围时，移除最早的点
        if(self.series_1.count()>299):
            self.series_1.removePoints(0,self.series_1.count()-299)

        #if(self.series_2.count()>299):
        #   self.series_2.removePoints(0,self.series_2.count()-299)
        # 产生随即数
        #yint = random.randint(0,1500)
        # 添加数据到曲线末端
        self.series_1.append(bjtime.toMSecsSinceEpoch(),yint)

        #yint = random.randint(0, 1500)
        # 添加数据到曲线末端
        #self.series_2.append(bjtime.toMSecsSinceEpoch(), yint)
        
        #自适应Y轴的显示范围
        if(self.series_1.count() > 0):
            ymin = 0
            ymax = 0
            for i in range(self.series_1.count()):
                if(self.series_1.at(i).y() > ymax) :
                    ymax = self.series_1.at(i).y()
                if(self.series_1.at(i).y() < ymin) :
                    ymin = self.series_1.at(i).y()
            self.setRange(ymin*0.8,ymax*1.2)
        
    def clear(self):
        self.series_1.clear()
        return None
    
    def setTitle(self, title):
        self.series_1.setName(title)
        self.vlaxisY.setTitleText(title)
        
    def setRange(self, min, max):
        self.vlaxisY.setMin(min)
        self.vlaxisY.setMax(max)


class Window(QMainWindow):

    def __init__(self):
        super().__init__()
  
        # setting title
        self.setWindowTitle("温度采集系统")

        # setting geometry
        self.setGeometry(0, 0, 1920, 1080)
  
        self.data_queue = queue.Queue()
        
        self.data_save = []
        
        self.threads = []
        
        self.starting = False

        # calling method
        self.UiComponents()
        
        icon = QIcon()
        icon.addPixmap(QPixmap("images/temperature-2010493-1697187-1.ico"),QIcon.Normal, QIcon.Off)
        self.setWindowIcon(icon)

        # showing all the widgets
        self.show()
        
    def request(self) :
        
        return None
  
    def recv_cb(self, q) :
        
        # 打开 COM17，将波特率配置为115200, 读超时时间为1秒
        print(self.port.currentText())
        ser = serial.Serial(port=self.port.currentText(), baudrate=9600, bytesize=8,parity=serial.PARITY_NONE, stopbits=serial.STOPBITS_ONE, timeout=1)

        # 读取串口输入信息并输出。
        while self.starting:
            if self.requestcb.isChecked() :
                req = bytearray(8)
                req[0] = 0x5A
                req[1] = 0xA5
                req[2] = 8
                req[3] = 7
                req[4] = 0xBA
                req[5] = 1
                req[6] = 0xAA
                req[7] = 0x55
                ser.write(req)
                ser.flush()
            #time.sleep(100)
            com_input = ser.read(100)
            if com_input:  # 如果读取结果非空，则输出
                _start_idx = -1
                _end_idx = 0
                print("Origin data:")
                print(com_input)
                
                for i in range(len(com_input)-1) :
                    if com_input[i] == 0x5A and com_input[i+1] == 0xA5 :
                        if len(com_input)-i >= 17 and com_input[15] == 0xAA and com_input[16] == 0x55 : 
                            _start_idx = i
                    if _start_idx >= 0:
                        _data = bytearray(17)
                        for j in range(17):
                            _data[j] = com_input[j+_start_idx]
                        print("Data:")
                        print(_data)
                            
                        temp1 = struct.unpack_from("<f", _data, 5)
                        temp2 = struct.unpack_from(">h", _data, 9)
                        k1 = _data[11]
                        k2 = _data[12]
                        rout = struct.unpack_from(">h", _data, 13)
                        print("temp1:")
                        print(temp1[0])
                        print("temp2:")
                        print(temp2[0])
                        print("K1:")
                        print(k1)
                        print("K2:")
                        print(k2)
                        print("PID:")
                        print(rout)
                        self.data_queue.put((temp1[0],temp2[0], k1, k2, rout[0]))
                        self.data_save.append((temp1[0],temp2[0], k1, k2, rout[0]))
                        
                        i = _start_idx+17
                        _start_idx = -1


        ser.close()
        
    # start data receiving thread
    def recv_start(self) :
        self.starting = True
        t = threading.Thread(target=self.recv_cb, args=(self.data_queue,))
        t.start()
        self.threads.append(t)     
        self.startbtn.setText("停止采集")
        
        return None
    
        # start data receiving thread
    def recv_stop(self) :
        self.starting = False
        for t in self.threads :
            t.join()
        self.threads.clear()
        self.view1.clear()
        self.view2.clear()
        self.data_save.clear()
        self.starting = False
        self.startbtn.setText("开始采集")
        
        return None
    
    def start(self):
        if not self.starting :
            self.recv_start() 
        else :
            self.recv_stop()
            
    def closeEvent(self, event):
        result = QtWidgets.QMessageBox.question(self, "标题", "亲，你确定退出监控程序！！！'_'", QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        if(result == QtWidgets.QMessageBox.Yes):
            if self.starting :
                self.recv_stop()
            event.accept()
        else:
            event.ignore()
            
    def append_data_to_excel(self, excel_name, sheet_name, new_columns, data):
        with pd.ExcelWriter(excel_name) as writer:
            _columns = []
            #for k, v in data.items():
             #   _columns.append(k)
            df = pd.DataFrame(data, index= None, columns=new_columns)
            df_source = None
            if os.path.exists(excel_name):
                df_source = pd.DataFrame(pd.read_excel(excel_name, sheet_name=sheet_name))
            if df_source is not None:
                df_dest = df_source.append(df)
            else:
                df_dest = pd.Dataframe(df)
            df_dest.to_excel(writer, sheet_name=sheet_name, index = True, columns=new_columns)   
            

    def append_dataframe_to_excel(self, df, excel_file):
        #df = pd.DataFrame(self.data_save, columns = ['实时温度值', 'AD值', 'K1', 'K2', 'PID输出'])
        print(df)
        # render dataframe as html
        sheet_name = "温度值记录"
        startrow = 0
        if not os.path.exists(excel_file):      
            writer = pd.ExcelWriter(excel_file,engine='openpyxl')
            df.to_excel(writer,sheet_name = sheet_name, startrow=startrow,index=False)
            writer.save()
            writer.close()
        else :
            '''
            wb = load_workbook(self.filename)
            writer = pd.ExcelWriter(self.filename,engine='openpyxl')
            writer.book = wb        
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row
            df.to_excel(writer,sheet_name = sheet_name, startrow=startrow)
            writer.save()
            writer.close()
            '''

            # appending the data of df after the existing data
            with pd.ExcelWriter(self.filename,mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer:
                df.to_excel(writer, sheet_name=sheet_name,header=None, startrow=writer.sheets[sheet_name].max_row,index=False)    
            
        if startrow > 10000:
            # saving file
            filename = self.generatefilename()
            os.rename(self.filename, filename)
    
    def generatefilename(self) :
        # get current date and time
        current_datetime = datetime.now()
        print("Current date & time : ", current_datetime)
        
        # convert datetime obj to string
        str_current_datetime = datetime.now().strftime("%Y%m%d%H%M%S")
  
        # create a file object along with extension
        file_name = str_current_datetime+".xlsx"
        
        return file_name
    
    def add_trend_chart_to_excel(self, filename):      
         with pd.ExcelWriter(filename,mode="a",engine="openpyxl",if_sheet_exists="overlay") as writer:
                #df.to_excel(writer, sheet_name=sheet_name,header=None, startrow=writer.sheets[sheet_name].max_row,index=False)  
                # 使用XlsxWriter作为引擎创建Excel编写器。              
                sheet_name = "温度值记录"
                startrow=writer.sheets[sheet_name].max_row
                # 将数据框转换为XlsxWriter Excel对象。
                #temp_data = pd.read_excel(writer, sheet_name = sheet_name, engine='xlsxreader')

                # 获取xlsxwriter工作簿和工作表对象。
                workbook  = writer.book
                worksheet = workbook.active
                workbook.title = sheet_name
              
                values = Reference(workbook, min_col= 2, min_row=2, max_col = 3, max_row = startrow)
                x_values = Reference(workbook, range_string=sheet_name+"!A2:A"+str(startrow))
                
                #initialize LineChart object
                chart = LineChart()
                #add data to the LineChart object
                chart.add_data(values, titles_from_data = True)
                #set x-axis
                chart.set_categories(x_values)

                ##cosmetics
                chart.title = '温度控制趋势图'
                chart.x_axis.title = '序号'
                chart.y_axis.title = '实时温度值'
                chart.legend.position = 'b'

                # 将图表插入工作表，指定图表的位置
                worksheet.insert_chart('D2', chart)
                writer.save();

    def add_trend_chart(self, file_name) :
        # read the data
        excel_file_data = load_workbook ( file_name )
        excel_file_data.sheetnames

        sheet_values = excel_file_data['温度值记录']
        #print(" *** One of the value from the sheet is - { sheet_values [ 'A2' ] . value } - { sheet_values [ 'B2' ] . value } " )
        '''
        for row in sheet_values :
            for cell in row :
                print(cell.value ) # barchart creation from openpyxl.chart
        '''
            
        chart = LineChart()
        # Fill the basic information like chart title,..
        chart . title = "温度趋势曲线"
        chart . y_axis . title = '实时温度值'
        chart . x_axis . title = '时间'
        new_workbook = excel_file_data.active
        # Now we will create a reference to the data and append the data to the chart.
        data = Reference(new_workbook , min_col= 1, min_row=1, max_col = 2, max_row = new_workbook.max_row)
        #x_values = Reference(new_workbook, range_string="温度值记录!A2:A698")
        x_values = Reference(new_workbook , min_col= 1, min_row=2, max_col = 1, max_row = new_workbook.max_row)
        y_values = Reference(new_workbook, range_string="温度值记录!B2:B100")
        chart.add_data (data , titles_from_data = True )
        #chart.set_categories(x_values)
        # Finally, Add the chart to the sheet and save the file.
        new_workbook.add_chart ( chart , "G2" )
        excel_file_data.save(file_name)
    
            
    def export(self):       
        if(len(self.data_save) >= 10) :
            df = pd.DataFrame(self.data_save, columns = ['实时温度值', 'AD值', 'K1', 'K2', 'PID输出'] )
            self.append_dataframe_to_excel(df, self.filename)
            self.data_save.clear()
            
        # saving file
        if os.path.exists(self.filename):
            filename = self.generatefilename()
            os.rename(self.filename, filename)
            
            #filename = "20220802223727.xlsx"
            self.add_trend_chart(filename)
                
            print('DataFrame is written successfully to the Excel File.')
            os.system('start '+ filename)
        return None
    
    def setting(self):
        if self.starting :
            self.recv_stop()
        self.MainDialog = QDialog()  
        self.myDialog = SettingDialog.Ui_SettingDialog()     
        self.myDialog.setupUi(self.MainDialog)
        
        self.myDialog.setCom(self.port.currentText())
        self.MainDialog.show()
        '''
        self.sd = Setting_Dialog()
        self.sd.show()
        '''
        return None
    
    def addTableRow(self, row_data):
        row = self.tableview.rowCount()
        if row > 10 :
            self.tableview.clearContents()
            row = 0
        self.tableview.setRowCount(row+1)
        col = 0
        for item in row_data:
            cell = QTableWidgetItem(str(item))
            self.tableview.setItem(row, col, cell)
            col += 1
        self.tableview.scrollToBottom()

    # method for components
    def UiComponents(self):

        # creating a widget object
        widget = QWidget()

        # creating a push button object
        self.startbtn = QPushButton('开始采集')
        self.startbtn.clicked.connect(self.start)

        # creating a line edit widget
        self.port = QComboBox(self)
        self.port.addItems(["COM1", "COM2","COM3","COM4", "COM5", "COM6", "COM7", "COM8", "COM9", "COM10"])
        self.port.setCurrentIndex(5)

        # creating a check box widget
        self.exportbtn = QPushButton("数据导出")
        self.exportbtn.clicked.connect(self.export)
        self.exportbtn.height = self.startbtn.height
        
        # creating a check box widget
        self.settingbtn = QPushButton("设置")
        self.settingbtn.clicked.connect(self.setting)
        self.settingbtn.height = self.startbtn.height
        
        # creating a check box widget
        self.requestcb = QCheckBox("发送请求")
        self.requestcb.height = self.startbtn.height
        self.requestcb.setChecked(True)

        # creating a plot window
        #plot = pg.plot()

        # create list for y-axis
        self.y1 = [5, 5, 7, 10, 3, 8, 9, 1, 6, 2]

        # create horizontal list i.e x-axis
        self.x = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]

        # create pyqt5graph bar graph item
        # with width = 0.6
        # with bar colors = green
        self.bargraph = pg.PlotDataItem(x = self.x, y = self.y1, width = 0.6, brush ='g')

        # add item to plot window
        # adding bargraph item to the plot window
        # plot.addItem(self.bargraph)
        self.view1 = ChartView()
        self.view1.setTitle("实时温度值")
        self.view1.setRange(0,1000)
        self.view1.setMinimumHeight(400)

        self.view2 = ChartView()
        self.view2.setTitle("AD值")
        self.view2.setRange(0,5000)
        self.view2.setMinimumHeight(400)
        
        self.tableview = QTableWidget(self)
        self.tableview.setColumnCount(5)
        self.tableview.setHorizontalHeaderLabels(['实时温度值', 'AD值', 'K1', 'K2', 'PID输出'])
        self.tableview.setMinimumHeight(300)
        # Creating a grid layout
        layout = QGridLayout()

        # setting this layout to the widget
        widget.setLayout(layout)

        # adding widgets in the layout in their proper positions
        # button goes in upper-left
        layout.addWidget(self.startbtn, 0, 0)

        # text edit goes in middle-left
        layout.addWidget(self.port, 0, 1)

        # export widget goes in bottom-left
        layout.addWidget(self.exportbtn, 0, 2)
        
        # setting widget goes in bottom-left
        layout.addWidget(self.settingbtn, 0, 3)
        
        # setting widget goes in bottom-left
        layout.addWidget(self.requestcb, 0, 4)
        
        self.k1le = QtWidgets.QLineEdit(self)
        self.k1le.setObjectName("k1le")
        layout.addWidget(self.k1le, 0, 5)
        
        self.k2le = QtWidgets.QLineEdit(self)
        self.k2le.setObjectName("k2le")
        layout.addWidget(self.k2le, 0, 6)
        
        self.pidle = QtWidgets.QLineEdit(self)
        self.pidle.setObjectName("pidle")
        layout.addWidget(self.pidle, 0, 7)
        

        # plot window goes on right side, spanning 3 rows
        layout.addWidget(self.view1, 1, 0, 3, 8)

        layout.addWidget(self.view2, 4, 0, 3, 8)
        
        layout.addWidget(self.tableview, 7, 0, 10, 8)

        # setting this widget as central widget of the main window
        self.setCentralWidget(widget)

        # Create timer object
        self.idx = 0
        self.timer = QTimer(self)
        # Add a method with the timer
        self.timer.timeout.connect(self.Refresh)
        self.timer.start(100)
        
        self.filename = 'output.xlsx'
        #self.recv_start()

    def Refresh(self):
        '''
        self.idx +=1
        self.x.append(self.idx)
        self.y1.append(self.idx*self.idx)
        self.bargraph = pg.PlotDataItem(x=self.x, y=self.y1, width=0.6, brush='g')
        '''
        while not self.data_queue.empty() :
            elem = self.data_queue.get()
            print(elem)
            self.view1.drawLine(elem[0])
            self.view2.drawLine(elem[1])
            self.k1le.setText(str(elem[2]))
            self.k2le.setText(str(elem[3]))
            self.pidle.setText(str(elem[4]))
            self.addTableRow(elem)
            
        if(len(self.data_save) >= 10) :
            df = pd.DataFrame(self.data_save, columns = ['实时温度值', 'AD值', 'K1', 'K2', 'PID输出'])
            self.append_dataframe_to_excel(df, self.filename)
            self.data_save.clear()
        
            

# create pyqt5 app
App = QApplication(sys.argv)

# create the instance of our Window
window = Window()

# start the app
sys.exit(App.exec())
